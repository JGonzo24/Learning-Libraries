import os
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "AppleTV MAC.xlsx"

def get_user_input():
    print("\n--- Enter Apple TV Info (or type 'q' to quit) ---")
    building = input("Building number: ").strip()
    if building.lower() in {"q", "quit"}:
        return None

    room = input("Room number: ").strip()
    if room.lower() in {"q", "quit"}:
        return None

    serial = input("Serial number: ").strip()
    if serial.lower() in {"q", "quit"}:
        return None

    mac = input("MAC address: ").strip()
    if mac.lower() in {"q", "quit"}:
        return None

    return building, room, serial, mac

def init_workbook(filename):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Devices"
        ws.append(["Building", "Room", "Serial Number", "MAC Address"])
        wb.save(filename)

def append_to_excel(filename, data):
    wb = load_workbook(filename)
    ws = wb["Devices"]
    ws.append(data)
    wb.save(filename)
    print(f"✅ Saved entry to {filename}")

def main():
    init_workbook(EXCEL_FILE)

    while True:
        entry = get_user_input()
        if entry is None:
            print("👋 Exiting.")
            break
        append_to_excel(EXCEL_FILE, entry)

if __name__ == "__main__":
    main()
